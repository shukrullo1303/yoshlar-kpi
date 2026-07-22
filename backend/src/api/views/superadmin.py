import calendar
import io
import os
import zipfile
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from django.contrib.auth.models import User
from django.db.models import Sum
from django.http import HttpResponse
from rest_framework import permissions
from rest_framework.response import Response
from rest_framework.views import APIView

from src.core.models import KPIDirection, KPITask, Profile, TaskAttachment


class IsSuperAdmin(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and request.user.is_superuser)


class SuperAdminUserListView(APIView):
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        profiles = Profile.objects.select_related('user').order_by('mahalla_name')
        return Response([{
            'id': p.id,
            'user_id': p.user.id,
            'username': p.user.username,
            'first_name': p.user.first_name,
            'last_name': p.user.last_name,
            'full_name': p.user.get_full_name(),
            'mahalla_name': p.mahalla_name,
            'district': p.district,
            'is_active': p.user.is_active,
            'is_staff': p.user.is_staff,
            'is_hokim': p.is_hokim,
        } for p in profiles])

    def post(self, request):
        username     = request.data.get('username', '').strip()
        password     = request.data.get('password', '').strip()
        mahalla_name = request.data.get('mahalla_name', '').strip()
        first_name   = request.data.get('first_name', '').strip()
        last_name    = request.data.get('last_name', '').strip()
        district     = request.data.get('district', '').strip()
        is_hokim     = bool(request.data.get('is_hokim', False))

        if not username or not password:
            return Response({'error': 'username va password majburiy'}, status=400)
        if not is_hokim and not mahalla_name:
            return Response({'error': 'MFY yetakchisi uchun mahalla_name majburiy'}, status=400)
        if User.objects.filter(username=username).exists():
            return Response({'error': 'Bu username allaqachon mavjud'}, status=400)

        user = User.objects.create_user(
            username=username, password=password,
            first_name=first_name, last_name=last_name,
        )
        profile = Profile.objects.create(
            user=user, mahalla_name=mahalla_name, district=district, is_hokim=is_hokim,
        )
        return Response({'id': profile.id, 'username': username}, status=201)


class SuperAdminUserDetailView(APIView):
    permission_classes = [IsSuperAdmin]

    def patch(self, request, pk):
        try:
            profile = Profile.objects.select_related('user').get(pk=pk)
        except Profile.DoesNotExist:
            return Response({'error': 'Topilmadi'}, status=404)

        user = profile.user
        d = request.data

        if 'username' in d:
            new_un = d['username'].strip()
            if new_un and new_un != user.username:
                if User.objects.filter(username=new_un).exclude(pk=user.pk).exists():
                    return Response({'error': 'Bu username band'}, status=400)
                user.username = new_un

        if 'password' in d and d['password'].strip():
            user.set_password(d['password'].strip())

        if 'first_name' in d:
            user.first_name = d['first_name'].strip()
        if 'last_name' in d:
            user.last_name = d['last_name'].strip()
        if 'mahalla_name' in d:
            profile.mahalla_name = d['mahalla_name'].strip()
        if 'district' in d:
            profile.district = d['district'].strip()
        if 'is_active' in d:
            user.is_active = bool(d['is_active'])
        if 'is_hokim' in d:
            profile.is_hokim = bool(d['is_hokim'])

        user.save()
        profile.save()
        return Response({'ok': True})


class SuperAdminDirectionView(APIView):
    """Superadmin — yo'nalishlarni ko'rish, qo'shish, tahrirlash."""
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        dirs = KPIDirection.objects.all().order_by('order')
        return Response([{
            'id': d.id,
            'key': d.key,
            'label': d.label,
            'max_score': d.max_score,
            'order': d.order,
            'admin_scored': d.admin_scored,
            'is_uploadable': d.is_uploadable,
            'is_active': d.is_active,
            'default_target': d.default_target,
            'info': d.info,
            'how': d.how,
        } for d in dirs])

    def post(self, request):
        key   = request.data.get('key', '').strip()
        label = request.data.get('label', '').strip()
        try:
            max_score = int(request.data.get('max_score', 0))
        except (TypeError, ValueError):
            return Response({'error': 'max_score musbat son bo\'lishi kerak'}, status=400)

        if not key or not label or max_score < 1:
            return Response({'error': 'key, label va max_score (≥1) majburiy'}, status=400)
        if KPIDirection.objects.filter(key=key).exists():
            return Response({'error': 'Bu key allaqachon mavjud'}, status=400)

        order         = int(request.data.get('order', 0))
        admin_scored  = bool(request.data.get('admin_scored', False))
        is_uploadable = bool(request.data.get('is_uploadable', True))
        is_active     = bool(request.data.get('is_active', True))
        default_target = int(request.data.get('default_target', 0))
        info = request.data.get('info', '')
        how  = request.data.get('how', '')

        d = KPIDirection.objects.create(
            key=key, label=label, max_score=max_score, order=order,
            admin_scored=admin_scored, is_uploadable=is_uploadable,
            is_active=is_active, default_target=default_target,
            info=info, how=how,
        )
        return Response({'id': d.id, 'key': d.key}, status=201)


class SuperAdminDirectionDetailView(APIView):
    """Superadmin — bitta yo'nalishni tahrirlash."""
    permission_classes = [IsSuperAdmin]

    def patch(self, request, pk):
        try:
            d = KPIDirection.objects.get(pk=pk)
        except KPIDirection.DoesNotExist:
            return Response({'error': 'Topilmadi'}, status=404)

        data = request.data
        if 'label' in data:
            d.label = data['label'].strip()
        if 'max_score' in data:
            try:
                d.max_score = int(data['max_score'])
            except (TypeError, ValueError):
                return Response({'error': 'max_score son bo\'lishi kerak'}, status=400)
        if 'order' in data:
            d.order = int(data['order'])
        if 'admin_scored' in data:
            d.admin_scored = bool(data['admin_scored'])
        if 'is_uploadable' in data:
            d.is_uploadable = bool(data['is_uploadable'])
        if 'is_active' in data:
            d.is_active = bool(data['is_active'])
        if 'default_target' in data:
            d.default_target = int(data['default_target'])
        if 'info' in data:
            d.info = data['info']
        if 'how' in data:
            d.how = data['how']

        d.save()
        return Response({'ok': True})


class SuperAdminScoreView(APIView):
    """Superadmin — foydalanuvchi ballarini ko'rish va o'zgartirish."""
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        month_param = request.query_params.get('month')
        if month_param:
            try:
                m = date.fromisoformat(month_param)
                month = f'{m.year}-{m.month:02d}-01'
            except ValueError:
                month = None
        if not month_param or not month:
            d = date.today()
            month = f'{d.year}-{d.month:02d}-01'

        directions = list(KPIDirection.objects.filter(is_active=True).order_by('order'))
        profiles   = Profile.objects.select_related('user').order_by('mahalla_name')

        m_date  = date.fromisoformat(month)
        m_last  = date(m_date.year, m_date.month,
                       calendar.monthrange(m_date.year, m_date.month)[1])

        rows = []
        for p in profiles:
            scores = []
            for dir_ in directions:
                if dir_.key == '1_ijro':
                    qs = KPITask.objects.filter(
                        leader=p, direction=dir_.key, status='yashil',
                        month__gte=month, month__lte=str(m_last),
                    )
                else:
                    qs = KPITask.objects.filter(
                        leader=p, direction=dir_.key, status='yashil', month=month,
                    )
                total = min(float(qs.aggregate(Sum('score'))['score__sum'] or 0), dir_.max_score)
                scores.append({'direction': dir_.key, 'score': round(total, 2)})
            rows.append({
                'profile_id': p.id,
                'mahalla_name': p.mahalla_name,
                'full_name': p.user.get_full_name(),
                'scores': scores,
            })

        return Response({
            'month': month,
            'directions': [{'key': d.key, 'label': d.label, 'max_score': d.max_score}
                           for d in directions],
            'rows': rows,
        })

    def post(self, request):
        profile_id    = request.data.get('profile_id')
        direction_key = request.data.get('direction')
        month         = request.data.get('month')
        score         = request.data.get('score')

        if profile_id is None or not direction_key or not month or score is None:
            return Response({'error': 'profile_id, direction, month, score majburiy'}, status=400)

        try:
            profile   = Profile.objects.get(pk=profile_id)
            direction = KPIDirection.objects.get(key=direction_key)
        except (Profile.DoesNotExist, KPIDirection.DoesNotExist):
            return Response({'error': "Foydalanuvchi yoki yo'nalish topilmadi"}, status=404)

        score = float(score)
        if score < 0 or score > direction.max_score:
            return Response({'error': f"Ball 0–{direction.max_score} bo'lishi kerak"}, status=400)

        # Normalize month to YYYY-MM-01
        try:
            m = date.fromisoformat(month)
            month = f'{m.year}-{m.month:02d}-01'
        except ValueError:
            return Response({'error': 'month format: YYYY-MM-01'}, status=400)

        # Keep or create one approved task, delete duplicates
        tasks = KPITask.objects.filter(leader=profile, direction=direction_key, month=month)
        yashil = tasks.filter(status='yashil')

        if yashil.exists():
            first = yashil.first()
            yashil.exclude(pk=first.pk).delete()
            first.score = score
            first.save()
        else:
            # Reject any pending tasks and create a fresh approved one
            tasks.filter(status='sariq').update(status='qizil')
            KPITask.objects.create(
                leader=profile, direction=direction_key,
                month=month, status='yashil', score=score,
            )

        return Response({'ok': True, 'score': score})


class SuperAdminMediaExportView(APIView):
    """
    GET  — barcha media fayllarni ZIP + Excel metadata sifatida yuklab olish
    DELETE — barcha media fayllarni diskdan va DB dan o'chirish
    """
    permission_classes = [IsSuperAdmin]

    _STATUS = {'sariq': 'Kutilmoqda', 'yashil': 'Tasdiqlangan', 'qizil': 'Rad etilgan'}

    def _direction_label(self, key):
        try:
            return KPIDirection.objects.get(key=key).label
        except KPIDirection.DoesNotExist:
            return key

    def get(self, request):
        attachments = (
            TaskAttachment.objects
            .select_related('task', 'task__leader', 'task__leader__user')
            .order_by('task__leader__mahalla_name', 'task__direction', 'task__month')
        )

        dir_labels = {d.key: d.label for d in KPIDirection.objects.all()}

        # ── Excel ──────────────────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Media fayllar'

        headers = ['#', 'MFY nomi', 'Ism Familiya', "Yo'nalish", 'Oy', 'Ball', 'Holat',
                   'Yuklangan sana', 'Fayl turi', 'Fayl nomi']
        ws.append(headers)

        hdr_fill = PatternFill(fill_type='solid', fgColor='1E293B')
        hdr_font = Font(bold=True, color='FFFFFF')
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for i, att in enumerate(attachments, 1):
            task    = att.task
            profile = task.leader
            month_str = str(task.month)[:7]
            ws.append([
                i,
                f"{profile.mahalla_name} MFY",
                profile.user.get_full_name() or profile.user.username,
                dir_labels.get(task.direction, task.direction),
                month_str,
                task.score or 0,
                self._STATUS.get(task.status, task.status),
                att.created_at.strftime('%Y-%m-%d %H:%M') if att.created_at else '',
                'Rasm' if att.is_image else 'PDF/Fayl',
                os.path.basename(att.file.name) if att.file else '',
            ])

        col_widths = [4, 22, 20, 22, 10, 8, 14, 18, 10, 30]
        for col_cells, w in zip(ws.columns, col_widths):
            ws.column_dimensions[col_cells[0].column_letter].width = w

        excel_buf = io.BytesIO()
        wb.save(excel_buf)
        excel_buf.seek(0)

        # ── ZIP ────────────────────────────────────────────────────────────────
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr('metadata.xlsx', excel_buf.read())
            for att in attachments:
                if att.file:
                    try:
                        if os.path.exists(att.file.path):
                            zf.write(att.file.path, att.file.name)
                    except Exception:
                        pass

        zip_buf.seek(0)
        filename = f"media-export-{date.today().isoformat()}.zip"
        response = HttpResponse(zip_buf.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def delete(self, request):
        attachments = TaskAttachment.objects.all()
        deleted_files = 0
        for att in attachments:
            if att.file:
                try:
                    if os.path.exists(att.file.path):
                        os.remove(att.file.path)
                        deleted_files += 1
                except Exception:
                    pass
        count = attachments.count()
        attachments.delete()
        return Response({'ok': True, 'deleted_records': count, 'deleted_files': deleted_files})
